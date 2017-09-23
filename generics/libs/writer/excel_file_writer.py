from openpyxl.workbook.workbook import Workbook

from engine.exceptions.br_exception import BRException
from generics.libs.writer.writter import Writter


class ExcelFileWriter(Writter):
    def __init__(self, data, header=[], file_path=None, *args, **kwargs):
        super(ExcelFileWriter, self).__init__(data=data, header=header,
                                              file_path=file_path, *args, **kwargs)

    def write(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.get_active_sheet()

        if not len(self.header):
            raise BRException("Header not provided")

        if self.data and len(self.header) != len(self.data[0]):
            raise BRException("Header and Data must be in same format")

        # Write headers
        row = 1
        col = 1
        for col_header in self.header:
            self.sheet.cell(row=row, column=col, value = col_header)
            col += 1

        row += 1
        # Now write data
        for data_row in self.data:
            col = 1
            for col_value in data_row:
                col_value = col_value if col_value else ''
                self.sheet.cell(row=row, column=col, value=col_value)
                col += 1
            row += 1
        if self.response:
            self.workbook.save(self.response)
        else:
            if self.file_name:
                self.workbook.save(self.file_name)

